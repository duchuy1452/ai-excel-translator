from io import BytesIO
import re
from typing import Callable, Dict, List, Tuple

from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.datavalidation import DataValidation

from constants import CHARS_PER_TOKEN, MAX_TOKENS_PER_REQUEST, TRANSLATION_DELIMITER
from translators import GeminiTranslator
from utils import logger


class ExcelProcessor:
    def __init__(self, translator: GeminiTranslator):
        self.translator = translator

    @staticmethod
    def is_formula(cell: Cell) -> bool:
        """Checks if a cell contains a formula."""
        return cell.data_type == "f" or (
            isinstance(cell.value, str) and cell.value.upper().startswith("=")
        )

    @staticmethod
    def is_translatable(cell: Cell) -> bool:
        """Checks if a cell's content should be translated."""
        return (
            isinstance(cell.value, str)
            and not ExcelProcessor.is_formula(cell)
            and cell.value.strip() != ""
        )

    @staticmethod
    def estimate_tokens(text: str) -> int:
        """Estimates the number of tokens in a text."""
        return len(text) // CHARS_PER_TOKEN

    def create_batches(self, texts: List[str]) -> List[List[str]]:
        """Creates batches of texts based on the token limit."""
        batches = []
        current_batch = []
        current_tokens = 0

        for text in texts:
            text_tokens = self.estimate_tokens(text + TRANSLATION_DELIMITER)
            if current_tokens + text_tokens > MAX_TOKENS_PER_REQUEST:
                if current_batch:
                    batches.append(current_batch)
                current_batch = [text]
                current_tokens = text_tokens
            else:
                current_batch.append(text)
                current_tokens += text_tokens

        if current_batch:
            batches.append(current_batch)
        return batches

    def _translate_sheet_names(
        self, workbook: load_workbook, target_lang: str, sheet_names: List[str]
    ):
        """Translates the names of the sheets in the workbook."""
        original_sheet_names = sheet_names
        if original_sheet_names:
            translated_sheet_names = self.translator.translate_batch(
                original_sheet_names, target_lang
            )
            for old_name, new_name in zip(original_sheet_names, translated_sheet_names):
                new_name = new_name.replace("/", " ")
                if old_name in workbook.sheetnames:
                    workbook[old_name].title = new_name

    def _collect_translatable_cells(
        self, sheet
    ) -> Tuple[List[str], List[Tuple[int, int]]]:
        """Collects translatable cell values and their positions."""
        cells_to_process = []
        cell_positions = []
        for row in sheet.iter_rows():
            for cell in row:
                if self.is_translatable(cell):
                    cells_to_process.append(cell.value)
                    cell_positions.append((cell.row, cell.column))
        return cells_to_process, cell_positions

    def _collect_cells_in_range(
        self, sheet, cell_range: str
    ) -> Tuple[List[str], List[Tuple[int, int]]]:
        cells_to_process = []
        cell_positions = []
        for row in sheet[cell_range]:
            for cell in row:
                if self.is_translatable(cell):
                    cells_to_process.append(cell.value)
                    cell_positions.append((cell.row, cell.column))

        return cells_to_process, cell_positions

    def _collect_hyperlinks(self, sheet) -> Dict[Tuple[int, int], str]:
        """Collects hyperlinks with display text."""
        hyperlinks_to_translate = {}
        for row in sheet.iter_rows():
            for cell in row:
                if cell.hyperlink and cell.hyperlink.display:
                    hyperlinks_to_translate[(cell.row, cell.column)] = (
                        cell.hyperlink.display
                    )
        return hyperlinks_to_translate

    def _collect_dropdowns(self, sheet) -> Dict[str, List[str]]:
        """Collects dropdown data validation lists."""
        dropdowns_to_translate = {}
        for validation in sheet.data_validations:
            if (
                isinstance(validation, DataValidation)
                and validation.type == "list"
                and validation.formula1
            ):
                dropdown_items = [
                    item
                    for item in re.split(
                        r",\s*|\s+", validation.formula1.replace('"', "")
                    )
                    if item
                ]
                if dropdown_items:
                    dropdowns_to_translate[validation.sqref] = dropdown_items
        return dropdowns_to_translate

    def _collect_translatable_elements(self, sheet) -> Tuple[
        List[str],
        List[Tuple[int, int]],
        Dict[Tuple[int, int], str],
        Dict[str, List[str]],
    ]:
        """Collects cells, hyperlinks, and dropdowns to be translated."""
        cells_to_process, cell_positions = self._collect_translatable_cells(sheet)
        hyperlinks_to_translate = self._collect_hyperlinks(sheet)
        dropdowns_to_translate = self._collect_dropdowns(sheet)

        return (
            cells_to_process,
            cell_positions,
            hyperlinks_to_translate,
            dropdowns_to_translate,
        )

    def _process_cell_batches(
        self, sheet, cells_to_process, cell_positions, target_lang, sheet_name
    ):
        """Processes and translates batches of cell content."""
        batches = self.create_batches(cells_to_process)
        cells_processed = 0
        no_batches = len(batches)
        processed_batch_count = 0
        for batch in batches:
            translations = self.translator.translate_batch(batch, target_lang)
            for j, translation in enumerate(translations):
                idx = cells_processed + j
                if idx < len(cell_positions):
                    row, col = cell_positions[idx]
                    sheet.cell(row=row, column=col).value = translation
            cells_processed += len(batch)
            processed_batch_count += 1
            logger.info(
                "Processing sheet %s: %s, %s/%s batches",
                sheet_name,
                str(int(processed_batch_count / no_batches * 100)) + "%",
                processed_batch_count,
                no_batches,
            )
        return cells_processed

    def _translate_hyperlinks(self, sheet, hyperlinks_to_translate, target_lang):
        """Translates the display text of hyperlinks."""
        if hyperlinks_to_translate:
            hyperlink_texts = list(hyperlinks_to_translate.values())
            translated_hyperlink_texts = self.translator.translate_batch(
                hyperlink_texts, target_lang
            )
            for (row, col), translated_text in zip(
                hyperlinks_to_translate.keys(), translated_hyperlink_texts
            ):
                sheet.cell(row=row, column=col).hyperlink.display = translated_text

    def _translate_dropdowns(self, sheet, dropdowns_to_translate, target_lang):
        """Translates the items in dropdown lists."""
        for sqref, items in dropdowns_to_translate.items():
            translated_items = self.translator.translate_batch(items, target_lang)
            translated_formula = f'"{",".join(translated_items)}"'
            for data_validation in sheet.data_validations:
                if (
                    isinstance(data_validation, DataValidation)
                    and data_validation.sqref == sqref
                    and data_validation.type == "list"
                ):
                    data_validation.formula1 = translated_formula

    def process_workbook(
        self,
        file_data: BytesIO,
        target_lang: str,
        selected_sheets: List[str],
        cell_range: str = "",
        progress_callback: Callable = None,
    ) -> BytesIO:
        workbook = load_workbook(file_data, data_only=False)

        sheets_to_process = [workbook[sheet_name] for sheet_name in selected_sheets]
        no_sheet = len(sheets_to_process)

        for sheet_idx, sheet in enumerate(sheets_to_process):
            if cell_range:
                cells_to_process, cell_positions = self._collect_cells_in_range(
                    sheet, cell_range
                )
            else:
                (
                    cells_to_process,
                    cell_positions,
                    hyperlinks_to_translate,
                    dropdowns_to_translate,
                ) = self._collect_translatable_elements(sheet)

            total_cells = len(cells_to_process)
            cells_processed = 0

            cells_processed += self._process_cell_batches(
                sheet, cells_to_process, cell_positions, target_lang, sheet.title
            )

            if not cell_range:
                self._translate_hyperlinks(sheet, hyperlinks_to_translate, target_lang)
                self._translate_dropdowns(sheet, dropdowns_to_translate, target_lang)

            if progress_callback:
                progress_callback(
                    sheet.title,
                    cells_processed,
                    total_cells,
                    sheet_idx,
                    no_sheet,
                )

        self._translate_sheet_names(workbook, target_lang, selected_sheets)

        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return output
